import openpyxl

def dataGenerator():
    # num_one=[['name1','passwd1'],['name2','passwd2'],['name3','passwd3'],['name4','passwd4'],['name5','passwd5']]
    # return num_one
    wk=openpyxl.load_workbook('C:\\Users\\meifute\\Desktop\\Data_user.xlsx')
    sh=wk['Sheet1']
    r=sh.max_row
    li=[]
    li_1=[]
    for i in range(1,r+1):
        li_1 = []
        un=sh.cell(i,1)
        up =sh.cell(1,2)
        li_1.insert(0,un.value)
        li_1.insert(1,up.value)
        li.insert(i-1,li_1)

    print(li)
    return li
